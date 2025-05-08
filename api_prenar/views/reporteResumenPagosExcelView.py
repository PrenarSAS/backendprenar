from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.db.models import Q
from api_prenar.models import Pago
from api_prenar.serializers.reporteResumenPagoSerializers import ReporteResumenPagoSerializers
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from collections import defaultdict

class downloadResumenPagoView(APIView):
    def get(self, request):
        try:
            # 1. Aplicar filtros
            order_code = request.query_params.get('order_code', None)
            start_date = request.query_params.get('start_date', None)
            end_date = request.query_params.get('end_date', None)
            name_cliente = request.query_params.get('name_cliente', None)

            filters = Q()
            if order_code:
                filters &= Q(id_pedido__order_code__icontains=order_code)
            if name_cliente:
                filters &= Q(id_pedido__id_client__name__icontains=name_cliente)
            if start_date and end_date:
                filters &= Q(payment_date__gte=start_date, payment_date__lte=end_date)
            elif start_date:
                filters &= Q(payment_date__gte=start_date)
            elif end_date:
                filters &= Q(payment_date__lte=end_date)

            pagos_qs = Pago.objects.filter(filters)
            if not pagos_qs.exists():
                return Response(
                    {"message": "No se encontraron pagos con los filtros especificados."},
                    status=status.HTTP_200_OK
                )

            # 2. Serializar los pagos
            serializer = ReporteResumenPagoSerializers(pagos_qs, many=True)
            pagos_serialized = serializer.data

            # 3. Agrupar pagos por pedido
            pagos_por_pedido = defaultdict(list)
            for pago_obj, pago_data in zip(pagos_qs, pagos_serialized):
                pedido = pago_obj.id_pedido
                if pedido:
                    pagos_por_pedido[pedido].append(pago_data)

            # 4. Crear el libro y la hoja de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pagos"

            # 5. Escribir el encabezado
            headers = [
                "N° Pedido",
                "Cliente",
                "Fecha Pedido",
                "$ Valor Total Pedido",
                "$ Total Pagado",
                "Recibo",
                "Fecha Pago",
                "$ Monto",
                "Método de Pago"
            ]
            ws.append(headers)
            # Poner en negrita el encabezado
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # 6. Llenar filas
            current_row = 2
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for pedido, lista_pagos in pagos_por_pedido.items():
                num_payments = len(lista_pagos)

                # Si un pedido no tuviera pagos, forzamos 1 para evitar errores
                if num_payments == 0:
                    num_payments = 1
                    lista_pagos = [{}]

                start_row = current_row
                end_row = current_row + num_payments - 1

                # Calcular el total pagado según tu lógica
                total_paid = pedido.total - pedido.outstanding_balance if pedido else 0

                # Si hay más de un pago, combinamos las columnas 1..5 (N° Pedido, Cliente, etc.)
                # para que aparezcan como una sola celda vertical
                if num_payments > 1:
                    for col in range(1, 6):
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col,
                            end_row=end_row,
                            end_column=col
                        )

                # Llenar los datos del pedido en la primera fila del bloque
                ws.cell(row=start_row, column=1).value = pedido.order_code if pedido else ""
                ws.cell(row=start_row, column=2).value = (
                    pedido.id_client.name if pedido and hasattr(pedido, "id_client") else ""
                )
                ws.cell(row=start_row, column=3).value = (
                    str(pedido.order_date) if pedido else ""
                )
                ws.cell(row=start_row, column=4).value = pedido.total if pedido else 0
                ws.cell(row=start_row, column=5).value = total_paid

                # Ahora, cada pago se escribe en una fila distinta, pero sólo en columnas 6..9
                for i, pago_data in enumerate(lista_pagos):
                    row_index = current_row + i
                    ws.cell(row=row_index, column=6).value = pago_data.get("receipt_number", "")
                    ws.cell(row=row_index, column=7).value = str(pago_data.get("payment_date", ""))
                    ws.cell(row=row_index, column=8).value = pago_data.get("amount", 0)
                    ws.cell(row=row_index, column=9).value = pago_data.get("payment_method", "")

                # Avanzar el current_row para el siguiente pedido
                current_row += num_payments

            # 7. Ajustar ancho de columnas
            max_col = ws.max_column
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                # Ajusta el ancho según convenga
                if col_num in [1, 2, 6, 9]:
                    ws.column_dimensions[col_letter].width = 20
                else:
                    ws.column_dimensions[col_letter].width = 18

            # 8. Dar formato a las columnas de valores monetarios
            #    Columnas 4 ($ Valor Total Pedido), 5 ($ Total Pagado), 8 (Monto)
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Columna 4 (Valor total pedido)
                row_cells[3].number_format = '#,##0.00'
                # Columna 5 (Total pagado)
                row_cells[4].number_format = '#,##0.00'
                # Columna 8 (Monto)
                row_cells[7].number_format = '#,##0.00'

            # Aplicar borde a todas las celdas
            for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=ws.max_column
            ):
                for cell in row:
                    cell.border = thin_border

            # 9. Retornar el archivo Excel
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="ReporteResumenPagos.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response(
                {"message": "Error al generar el archivo.", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )