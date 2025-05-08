from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from api_prenar.serializers.reportePedidosResumenSerializers import ReportePedidosResumenSerializer
from django.db.models import Q
from api_prenar.models import Pedido
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Font

class downloadPedidosResumenView(APIView):
    def get(self, request):
        try:
            # 1. Aplicar filtros
            order_code = request.query_params.get('order_code', None)
            start_date = request.query_params.get('start_date', None)
            end_date = request.query_params.get('end_date', None)
            name_cliente = request.query_params.get('name_cliente', None)

            from django.db.models import Q
            filters = Q()
            if order_code:
                filters &= Q(order_code__icontains=order_code)
            if name_cliente:
                filters &= Q(id_client__name__icontains=name_cliente)
            if start_date and end_date:
                filters &= Q(order_date__gte=start_date, order_date__lte=end_date)
            elif start_date:
                filters &= Q(order_date__gte=start_date)
            elif end_date:
                filters &= Q(order_date__lte=end_date)

            pedidos_qs = Pedido.objects.filter(filters).order_by('-order_date')
            if not pedidos_qs.exists():
                return Response(
                    {"message": "No se encontraron pedidos con los filtros especificados."},
                    status=status.HTTP_200_OK
                )

            # 2. Serializar los pedidos
            serializer = ReportePedidosResumenSerializer(pedidos_qs, many=True)
            pedidos = serializer.data  # Lista de dicts

            # 3. Crear el libro y la hoja de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pedidos"

            # 4. Escribir el encabezado
            headers = [
                "N° Pedido",
                "Cliente",
                "Fecha Pedido",
                "$ Valor Total Pedido",
                "$ Valor Pagado",
                "$ Valor Por Cobrar",
                "Producto",
                "Cantidad Solicitada",
                "Cantidad Despachada",
                "Cantidad a Producir / Despachar",
                "$ Valor unidades Pendientes a Producir / Despachar"
            ]
            ws.append(headers)
            # Aplicar negrita a cada celda del encabezado (fila 1)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # 5. Llenar las filas
            # Empezamos a escribir datos a partir de la fila 2 (la 1 es encabezado)
            current_row = 2  

            for pedido in pedidos:
                products = pedido.get("products", [])
                num_products = len(products)

                # Si el pedido no tiene productos, igual creamos una fila
                # (aunque sea con celdas vacías para los productos)
                if num_products == 0:
                    num_products = 1
                    products = [{}]

                # Combinar las celdas para las columnas que se repiten
                # (order_code, client_name, order_date, total, value_paid, outstanding_balance)
                # desde current_row hasta current_row + num_products - 1  
                start_row = current_row
                end_row = current_row + num_products - 1

                # Columnas fijas: 1..6 (A..F en Excel)
                # Combinar cada una de esas columnas si num_products > 1
                if num_products > 1:
                    for col in range(1, 7):
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col,
                            end_row=end_row,
                            end_column=col
                        )

                # Llenar los datos fijos en la primera fila del bloque
                ws.cell(row=start_row, column=1).value = pedido.get("order_code", "")
                ws.cell(row=start_row, column=2).value = pedido.get("client_name", "")
                ws.cell(row=start_row, column=3).value = pedido.get("order_date", "")
                ws.cell(row=start_row, column=4).value = pedido.get("total", 0)
                ws.cell(row=start_row, column=5).value = pedido.get("value_paid", 0)
                ws.cell(row=start_row, column=6).value = pedido.get("outstanding_balance", 0)

                # Ahora iteramos sobre los productos y los vamos poniendo en filas sucesivas
                for i, product in enumerate(products):
                    row_index = current_row + i
                    name_color = ""
                    cantidad_unidades = 0
                    cantidades_despachadas = 0
                    cantidad_faltante = 0

                    # Extraer datos del product (si existe)
                    if product:
                        name_color = f'{product.get("name","")} ({product.get("color","")})'
                        cantidad_unidades = product.get("cantidad_unidades", 0)
                        cantidades_despachadas = product.get("cantidades_despachadas", 0)
                        cantidad_faltante = cantidad_unidades - cantidades_despachadas
                        valor_unidades_pendientes = product.get("valor_unidades_pendientes",0)

                    # Colocar datos del producto en columnas G..J (7..10)
                    ws.cell(row=row_index, column=7).value = name_color
                    ws.cell(row=row_index, column=8).value = cantidad_unidades
                    ws.cell(row=row_index, column=9).value = cantidades_despachadas
                    ws.cell(row=row_index, column=10).value = cantidad_faltante
                    ws.cell(row=row_index, column=11). value = valor_unidades_pendientes

                # Avanzar el current_row para el siguiente pedido
                current_row += num_products
            # Ajustar ancho general de las columnas
            max_col = ws.max_column
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 55
            
            # Definir un borde fino para aplicar a las celdas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Aplicar borde a todas las celdas usadas
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # 6. Retornar el archivo Excel
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="ReporteResumenPedidos.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response(
                {"message": "Error al generar el archivo.", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )